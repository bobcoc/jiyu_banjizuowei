import xml.etree.ElementTree as ET
from openpyxl import Workbook
import openpyxl.utils
import glob
import os

def set_cell_dimensions(ws, char_count=12):
    # 设置列宽和行高为20个字符的宽度
    approx_char_size = 1  # 根据Excel的默认字体和大小进行定量
    column_width = char_count * approx_char_size
    row_height = char_count * 3  # 行高也设置为相同的大小
    
    # 设置所有列的宽度
    for col in range(1, ws.max_column + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        ws.column_dimensions[col_letter].width = column_width

    # 设置所有行的高度
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = row_height

def group_students_by_y(students, y_threshold=15):#高度相差不超过15视为同一排
    sorted_students = sorted(students, key=lambda student: student[2])
    groups = []
    current_group = []
    current_y = sorted_students[0][2]
    
    for student in sorted_students:
        if abs(student[2] - current_y) <= y_threshold:
            current_group.append(student)
        else:
            groups.append(current_group)
            current_group = [student]
            current_y = student[2]
            
    # 添加最后一组
    groups.append(current_group)
    return groups

# 枚举当前目录下的所有.cls文件
for cls_file in glob.glob('*.cls'):
    tree = ET.parse(cls_file)
    root = tree.getroot()

    students = []

    for student in root.findall('.//student'):  # 路径根据实际的XML结构调整
        name = student.find('name').text
        x_pos = int(student.find('posThumb').attrib['x'])
        y_pos = int(student.find('posThumb').attrib['y'])
        students.append((name, x_pos, y_pos))

    student_groups = group_students_by_y(students)

    # 创建工作簿和工作表
    wb = Workbook()
    ws = wb.active

    # 将学生填入表格
    excel_row = 1
    for group in student_groups:
        group.sort(key=lambda student: student[1])  # 按x排序
        excel_col = 1
        for name, _, _ in group:
            ws.cell(row=excel_row, column=excel_col, value=name)
            excel_col += 1
        excel_row += 1

    # 调整单元格大小，确保所有数据已填入
    set_cell_dimensions(ws, char_count=20)

    # 保存为同名的xlsx文件
    excel_filename = f"{os.path.splitext(cls_file)[0]}.xlsx"
    wb.save(excel_filename)